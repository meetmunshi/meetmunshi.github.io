from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Dict, Optional, Tuple
import uuid
from datetime import datetime, timezone
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="Resource Scheduling Board v2")
api_router = APIRouter(prefix="/api")


# ----------------- Models -----------------
class Person(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sn: Optional[int] = None
    name: str
    surname: Optional[str] = ""
    qualification: Optional[str] = ""
    employee_type: Optional[str] = ""
    mobile: Optional[str] = ""
    skills: Dict[str, bool] = {}


class LineDetail(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    detail: str
    line: str
    row_name: str = ""
    persons_required: int


class LineConfig(BaseModel):
    line: str
    priority: int = 5           # 1 = highest
    run_count: int = 1          # 1..N replicas of this line


class ScheduleRequest(BaseModel):
    date: str
    shift: str = "day"           # day | evening | night
    line_configs: List[LineConfig] = []
    absent_person_ids: List[str] = []
    overrides: Dict[str, List[str]] = {}   # key = f"{row_name}||{line}#{run}" -> person_ids
    unassigned_keys: List[str] = []        # keys the user explicitly cleared


class CellAssignment(BaseModel):
    row_name: str
    line: str
    run: int                      # 1..run_count
    line_key: str                 # e.g., "X-Smart" or "X-Smart #2"
    detail: str
    required: int
    assigned_person_ids: List[str]
    assigned_person_names: List[str]
    shortage: int


class Schedule(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    shift: str = "day"
    line_configs: List[LineConfig]
    absent_person_ids: List[str]
    assignments: List[CellAssignment]
    overrides: Dict[str, List[str]] = {}
    unassigned_keys: List[str] = []
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    total_required: int = 0
    total_assigned: int = 0
    total_shortage: int = 0


# ----------------- Excel parsing & seeding -----------------
def _parse_persons_sheet(ws) -> List[Person]:
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    skill_start_col = 9
    skill_headers = [str(h).strip() for h in headers[skill_start_col - 1:] if h]

    persons: List[Person] = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name:
            continue
        sn_v = ws.cell(row=r, column=1).value
        skills = {
            sk: (str(ws.cell(row=r, column=skill_start_col + i).value).strip().lower() == "yes"
                 if ws.cell(row=r, column=skill_start_col + i).value is not None else False)
            for i, sk in enumerate(skill_headers)
        }
        persons.append(Person(
            sn=int(sn_v) if isinstance(sn_v, (int, float)) else None,
            name=str(name).strip(),
            surname=str(ws.cell(row=r, column=3).value or "").strip(),
            qualification=str(ws.cell(row=r, column=5).value or "").strip(),
            employee_type=str(ws.cell(row=r, column=6).value or "").strip(),
            mobile=str(ws.cell(row=r, column=8).value or "").strip(),
            skills=skills,
        ))
    return persons


def _coerce_int(v) -> int:
    try:
        return int(v) if v is not None else 0
    except (ValueError, TypeError):
        return 0


def _parse_lines_sheet(ws) -> List[LineDetail]:
    details: List[LineDetail] = []
    for r in range(2, ws.max_row + 1):
        detail = ws.cell(row=r, column=2).value
        line = ws.cell(row=r, column=3).value
        if not detail or not line:
            continue
        row_name = ws.cell(row=r, column=5).value
        details.append(LineDetail(
            detail=str(detail).strip(),
            line=str(line).strip(),
            row_name=str(row_name).strip() if row_name else str(detail).strip(),
            persons_required=_coerce_int(ws.cell(row=r, column=4).value),
        ))
    return details


def parse_excel_bytes(content: bytes) -> Tuple[List[Person], List[LineDetail]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet_names = {s.lower(): s for s in wb.sheetnames}
    persons_sheet = sheet_names.get("person - skill") or wb.sheetnames[0]
    lines_sheet = sheet_names.get("assembly line") or wb.sheetnames[1]
    persons = _parse_persons_sheet(wb[persons_sheet])
    details = _parse_lines_sheet(wb[lines_sheet])
    return persons, details


async def seed_from_file_if_empty():
    persons_count = await db.persons.count_documents({})
    details_count = await db.line_details.count_documents({})
    if persons_count > 0 and details_count > 0:
        return
    seed_path = ROOT_DIR / "seed_data.xlsx"
    if not seed_path.exists():
        return
    with open(seed_path, "rb") as f:
        content = f.read()
    persons, details = parse_excel_bytes(content)
    await db.persons.delete_many({})
    await db.line_details.delete_many({})
    if persons:
        await db.persons.insert_many([p.model_dump() for p in persons])
    if details:
        await db.line_details.insert_many([d.model_dump() for d in details])
    logger.info(f"Seeded {len(persons)} persons and {len(details)} line details")


# ----------------- Endpoints -----------------
@api_router.get("/")
async def root():
    return {"message": "Resource Scheduling API v2"}


@api_router.get("/persons", response_model=List[Person])
async def list_persons():
    docs = await db.persons.find({}, {"_id": 0}).to_list(1000)
    docs.sort(key=lambda p: (p.get("name", "").lower(), p.get("surname", "").lower()))
    return [Person(**d) for d in docs]


@api_router.get("/lines")
async def list_lines():
    docs = await db.line_details.find({}, {"_id": 0}).to_list(1000)
    by_line: Dict[str, List[dict]] = {}
    order: List[str] = []
    row_names_set: List[str] = []
    for d in docs:
        line = d["line"]
        if line not in by_line:
            by_line[line] = []
            order.append(line)
        by_line[line].append(d)
        rn = d.get("row_name") or d["detail"]
        if rn not in row_names_set:
            row_names_set.append(rn)
    return {
        "lines": [{"line": l, "details": by_line[l]} for l in order],
        "row_names": row_names_set,
    }


@api_router.get("/details", response_model=List[LineDetail])
async def list_details():
    docs = await db.line_details.find({}, {"_id": 0}).to_list(1000)
    return [LineDetail(**d) for d in docs]


@api_router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls files supported")
    content = await file.read()
    try:
        persons, details = parse_excel_bytes(content)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse excel: {e}")
    if not persons or not details:
        raise HTTPException(400, "No data found")
    await db.persons.delete_many({})
    await db.line_details.delete_many({})
    await db.persons.insert_many([p.model_dump() for p in persons])
    await db.line_details.insert_many([d.model_dump() for d in details])
    return {"persons": len(persons), "details": len(details)}


def _person_full_name(p: dict) -> str:
    return f"{p.get('name','').strip()} {p.get('surname','').strip()}".strip()


def _names_from_ids(ids: List[str], person_by_id: Dict[str, dict]) -> List[str]:
    return [_person_full_name(person_by_id[pid]) for pid in ids if pid in person_by_id]


def _apply_override_picks(
    cell_key: str,
    overrides: Dict[str, List[str]],
    person_by_id: Dict[str, dict],
    absent_set: set,
) -> List[str]:
    """Return picks locked in by user overrides for this cell (skips absent/unknown)."""
    if cell_key not in overrides:
        return []
    return [
        pid for pid in overrides[cell_key]
        if pid in person_by_id and pid not in absent_set
    ]


def _select_eligible(
    detail_name: str,
    remaining: int,
    persons: List[dict],
    absent_set: set,
    used_globally: set,
    used_locally: List[str],
    person_total_skills: Dict[str, int],
) -> List[dict]:
    """Return specialist-first eligible candidates for a cell."""
    eligible = [
        p for p in persons
        if bool(p.get("skills", {}).get(detail_name))
        and p["id"] not in absent_set
        and p["id"] not in used_globally
        and p["id"] not in used_locally
    ]
    eligible.sort(key=lambda p: (person_total_skills.get(p["id"], 0), p.get("name", "")))
    return eligible[:remaining]


def _build_work_items(
    line_configs: List[LineConfig],
    details_by_line: Dict[str, List[dict]],
) -> List[Tuple[int, str, int, dict]]:
    sorted_configs = sorted(line_configs, key=lambda c: (c.priority, c.line))
    work: List[Tuple[int, str, int, dict]] = []
    for cfg in sorted_configs:
        for run_idx in range(1, max(1, cfg.run_count) + 1):
            for d in details_by_line.get(cfg.line, []):
                work.append((cfg.priority, cfg.line, run_idx, d))
    return work


def _generate_assignments(
    persons: List[dict],
    details: List[dict],
    line_configs: List[LineConfig],
    absent_ids: List[str],
    overrides: Dict[str, List[str]],
    unassigned_keys: List[str],
) -> List[CellAssignment]:
    persons = persons or []
    details = details or []
    overrides = overrides or {}
    unassigned_keys = unassigned_keys or []

    person_total_skills = {p["id"]: sum(1 for v in p.get("skills", {}).values() if v) for p in persons}
    person_by_id = {p["id"]: p for p in persons}
    absent_set = set(absent_ids)
    unassigned_set = set(unassigned_keys)

    details_by_line: Dict[str, List[dict]] = defaultdict(list)
    for d in details:
        details_by_line[d["line"]].append(d)

    work = _build_work_items(line_configs, details_by_line)

    # Lock in override people up-front so auto-fill won't reuse them
    assigned_person_ids: set = set()
    for pids in overrides.values():
        for pid in pids:
            if pid in person_by_id and pid not in absent_set:
                assigned_person_ids.add(pid)

    results: List[CellAssignment] = []
    for _priority, line, run_idx, d in work:
        detail_name = d["detail"]
        row_name = d.get("row_name") or detail_name
        required = d["persons_required"]
        line_key = f"{line}" if run_idx == 1 else f"{line} #{run_idx}"
        cell_key = f"{row_name}||{line_key}"

        picks = _apply_override_picks(cell_key, overrides, person_by_id, absent_set)

        if cell_key not in unassigned_set:
            remaining = required - len(picks)
            if remaining > 0:
                extras = _select_eligible(
                    detail_name, remaining, persons,
                    absent_set, assigned_person_ids, picks, person_total_skills,
                )
                picks.extend(c["id"] for c in extras)

        assigned_person_ids.update(picks)
        results.append(CellAssignment(
            row_name=row_name, line=line, run=run_idx, line_key=line_key,
            detail=detail_name, required=required,
            assigned_person_ids=picks,
            assigned_person_names=_names_from_ids(picks, person_by_id),
            shortage=max(0, required - len(picks)),
        ))

    return results


@api_router.post("/schedule", response_model=Schedule)
async def generate_schedule(req: ScheduleRequest):
    persons = await db.persons.find({}, {"_id": 0}).to_list(1000)
    details = await db.line_details.find({}, {"_id": 0}).to_list(1000)
    if not persons or not details:
        raise HTTPException(400, "No data seeded.")
    assignments = _generate_assignments(
        persons, details, req.line_configs, req.absent_person_ids,
        req.overrides, req.unassigned_keys,
    )
    total_req = sum(a.required for a in assignments)
    total_assigned = sum(len(a.assigned_person_ids) for a in assignments)
    total_short = sum(a.shortage for a in assignments)
    sched = Schedule(
        date=req.date, shift=req.shift, line_configs=req.line_configs,
        absent_person_ids=req.absent_person_ids,
        assignments=assignments, overrides=req.overrides,
        unassigned_keys=req.unassigned_keys,
        total_required=total_req, total_assigned=total_assigned, total_shortage=total_short,
    )
    doc = sched.model_dump()
    await db.schedules.replace_one({"date": req.date, "shift": req.shift}, doc, upsert=True)
    return sched


class AutoPlanRequest(BaseModel):
    date: str
    shift: str = "day"
    absent_person_ids: List[str] = []
    min_coverage: int = 80          # only include lines with at least this coverage %


def _skill_totals(persons: List[dict]) -> Dict[str, int]:
    return {p["id"]: sum(1 for v in p.get("skills", {}).values() if v) for p in persons}


def _group_details_by_line(details: List[dict]) -> Dict[str, List[dict]]:
    grouped: Dict[str, List[dict]] = defaultdict(list)
    for d in details:
        grouped[d["line"]].append(d)
    return grouped


def _pick_specialists(
    detail_name: str,
    limit: int,
    persons: List[dict],
    skill_totals: Dict[str, int],
    excluded_ids: set,
) -> List[dict]:
    """Return up to `limit` skilled candidates, specialist-first, excluding used/absent."""
    eligible = [
        p for p in persons
        if bool(p.get("skills", {}).get(detail_name)) and p["id"] not in excluded_ids
    ]
    eligible.sort(key=lambda p: (skill_totals.get(p["id"], 0), p.get("name", "")))
    return eligible[:limit]


def _simulate_line_fill(
    line_details: List[dict],
    persons: List[dict],
    skill_totals: Dict[str, int],
    absent_set: set,
    reserved: set,
) -> Tuple[int, int, set, List[dict]]:
    """Simulate assigning `line_details` from the free pool.
    Returns (total_required, total_filled, used_ids, per_detail_reports)."""
    line_used: set = set()
    total_req = 0
    total_fill = 0
    reports: List[dict] = []
    for d in line_details:
        excluded = absent_set | reserved | line_used
        picks = _pick_specialists(d["detail"], d["persons_required"], persons, skill_totals, excluded)
        for c in picks:
            line_used.add(c["id"])
        total_req += d["persons_required"]
        total_fill += len(picks)
        reports.append({"detail": d, "picks": picks})
    return total_req, total_fill, line_used, reports


@api_router.post("/schedule/auto-plan", response_model=Schedule)
async def auto_plan(req: AutoPlanRequest):
    """Given absentees, greedily pick lines that maximize headcount utilization."""
    persons = await db.persons.find({}, {"_id": 0}).to_list(1000)
    details = await db.line_details.find({}, {"_id": 0}).to_list(1000)
    if not persons or not details:
        raise HTTPException(400, "No data seeded.")

    absent_set = set(req.absent_person_ids)
    skill_totals = _skill_totals(persons)
    details_by_line = _group_details_by_line(details)
    all_lines = list(details_by_line.keys())

    selected: List[str] = []
    used: set = set()

    def score(line: str) -> Tuple[int, int, set]:
        total_req, total_fill, line_used, _ = _simulate_line_fill(
            details_by_line[line], persons, skill_totals, absent_set, used,
        )
        pct = 100 if total_req == 0 else round(total_fill * 100 / total_req)
        return pct, total_fill, line_used

    while True:
        best = None
        best_score = (-1, -1)
        best_used: set = set()
        for line in all_lines:
            if line in selected:
                continue
            pct, fill, line_used = score(line)
            if pct >= req.min_coverage and (pct, fill) > best_score:
                best_score = (pct, fill)
                best = line
                best_used = line_used
        if not best:
            break
        selected.append(best)
        used.update(best_used)

    if not selected:
        raise HTTPException(400, "Could not auto-plan any line with the current headcount.")

    line_configs = [LineConfig(line=l, priority=i + 1, run_count=1) for i, l in enumerate(selected)]
    gen_req = ScheduleRequest(
        date=req.date, shift=req.shift,
        line_configs=line_configs,
        absent_person_ids=req.absent_person_ids,
        overrides={}, unassigned_keys=[],
    )
    return await generate_schedule(gen_req)


def _collect_used_person_ids(sched_doc: dict) -> set:
    used: set = set()
    for a in sched_doc["assignments"]:
        for pid in a["assigned_person_ids"]:
            used.add(pid)
    return used


def _build_suggestion(
    line: str, dets: List[dict], persons: List[dict],
    skill_totals: Dict[str, int], absent_set: set, used_globally: set,
) -> dict:
    total_req, total_filled, _, reports = _simulate_line_fill(
        dets, persons, skill_totals, absent_set, used_globally,
    )
    cell_reports = []
    for r in reports:
        d = r["detail"]
        picks = r["picks"]
        cell_reports.append({
            "row_name": d.get("row_name") or d["detail"],
            "detail": d["detail"],
            "required": d["persons_required"],
            "eligible": len(picks),
            "assigned_names": [_person_full_name(p) for p in picks],
            "shortage": max(0, d["persons_required"] - len(picks)),
        })
    coverage = 1.0 if total_req == 0 else total_filled / total_req
    return {
        "line": line,
        "required": total_req,
        "fillable": total_filled,
        "coverage_pct": round(coverage * 100),
        "fully_covered": total_filled == total_req and total_req > 0,
        "cells": cell_reports,
    }


@api_router.get("/schedule/{date}/suggest-lines")
async def suggest_lines(date: str, shift: str = "day"):
    """Using the free pool (unassigned + not absent), suggest additional lines that could run today."""
    sched_doc = await db.schedules.find_one({"date": date, "shift": shift}, {"_id": 0})
    if not sched_doc:
        raise HTTPException(404, "Schedule not found")

    persons = await db.persons.find({}, {"_id": 0}).to_list(1000)
    details = await db.line_details.find({}, {"_id": 0}).to_list(1000)
    absent_set = set(sched_doc.get("absent_person_ids", []))
    used = _collect_used_person_ids(sched_doc)

    active_lines = {c["line"] for c in sched_doc["line_configs"]}
    free_pool = [p for p in persons if p["id"] not in absent_set and p["id"] not in used]
    skill_totals = _skill_totals(persons)

    # Candidate lines = not active
    details_by_line = _group_details_by_line([d for d in details if d["line"] not in active_lines])
    suggestions = [
        _build_suggestion(line, dets, persons, skill_totals, absent_set, used)
        for line, dets in details_by_line.items()
    ]
    suggestions.sort(key=lambda s: (-s["coverage_pct"], -s["fillable"], s["line"]))
    return {"free_pool_size": len(free_pool), "suggestions": suggestions}


@api_router.post("/schedule/{date}/fill-shortages", response_model=Schedule)
async def fill_shortages(date: str, shift: str = "day"):
    """Auto-assign best available free candidates to every shortage cell."""
    sched_doc = await db.schedules.find_one({"date": date, "shift": shift}, {"_id": 0})
    if not sched_doc:
        raise HTTPException(404, "Schedule not found")

    persons = await db.persons.find({}, {"_id": 0}).to_list(1000)
    skill_totals = _skill_totals(persons)
    absent_set = set(sched_doc.get("absent_person_ids", []))
    used = _collect_used_person_ids(sched_doc)

    overrides = dict(sched_doc.get("overrides", {}) or {})
    unassigned = set(sched_doc.get("unassigned_keys", []) or [])

    def scarcity(cell: dict) -> int:
        return sum(
            1 for p in persons
            if bool(p.get("skills", {}).get(cell["detail"])) and p["id"] not in absent_set
        )

    shortage_cells = sorted(
        [a for a in sched_doc["assignments"] if a["shortage"] > 0],
        key=scarcity,
    )

    for cell in shortage_cells:
        cell_key = f"{cell['row_name']}||{cell['line_key']}"
        excluded = absent_set | used
        new_picks = _pick_specialists(cell["detail"], cell["shortage"], persons, skill_totals, excluded)
        new_ids = [p["id"] for p in new_picks]
        if not new_ids:
            continue
        overrides[cell_key] = cell["assigned_person_ids"] + new_ids
        unassigned.discard(cell_key)
        used.update(new_ids)

    req = ScheduleRequest(
        date=date, shift=shift,
        line_configs=[LineConfig(**c) for c in sched_doc["line_configs"]],
        absent_person_ids=sched_doc["absent_person_ids"],
        overrides=overrides,
        unassigned_keys=list(unassigned),
    )
    return await generate_schedule(req)


@api_router.post("/schedule/{date}/adjust")
async def adjust_cell(date: str, payload: dict):
    """Manual adjustment: swap or unassign a person in a cell.
    payload = {shift, cell_key, person_ids: [list], action: 'set' | 'clear'}"""
    shift = payload.get("shift", "day")
    cell_key = payload.get("cell_key")
    action = payload.get("action", "set")
    person_ids = payload.get("person_ids", [])
    if not cell_key:
        raise HTTPException(400, "cell_key required")
    sched_doc = await db.schedules.find_one({"date": date, "shift": shift}, {"_id": 0})
    if not sched_doc:
        raise HTTPException(404, "Schedule not found")
    overrides = sched_doc.get("overrides", {}) or {}
    unassigned = set(sched_doc.get("unassigned_keys", []) or [])
    if action == "clear":
        overrides.pop(cell_key, None)
        unassigned.add(cell_key)
    else:
        overrides[cell_key] = person_ids
        unassigned.discard(cell_key)
    req = ScheduleRequest(
        date=date, shift=shift,
        line_configs=[LineConfig(**c) for c in sched_doc["line_configs"]],
        absent_person_ids=sched_doc["absent_person_ids"],
        overrides=overrides,
        unassigned_keys=list(unassigned),
    )
    return await generate_schedule(req)


@api_router.get("/schedule/{date}", response_model=Optional[Schedule])
async def get_schedule(date: str, shift: str = "day"):
    doc = await db.schedules.find_one({"date": date, "shift": shift}, {"_id": 0})
    if not doc:
        return None
    return Schedule(**doc)


@api_router.get("/schedules")
async def list_schedules():
    docs = await db.schedules.find({}, {"_id": 0, "assignments": 0}).sort("date", -1).to_list(500)
    return docs


@api_router.delete("/schedule/{date}")
async def delete_schedule(date: str, shift: str = "day"):
    r = await db.schedules.delete_one({"date": date, "shift": shift})
    return {"deleted": r.deleted_count}


@api_router.get("/analytics/shortage")
async def shortage_analytics(days: int = 30):
    docs = await db.schedules.find({}, {"_id": 0}).sort("date", -1).limit(days).to_list(days)
    by_detail: Dict[str, int] = defaultdict(int)
    by_line: Dict[str, int] = defaultdict(int)
    by_date: List[dict] = []
    for s in docs:
        total = 0
        for a in s.get("assignments", []):
            if a["shortage"] > 0:
                by_detail[a["detail"]] += a["shortage"]
                by_line[a["line"]] += a["shortage"]
                total += a["shortage"]
        by_date.append({"date": s["date"], "shift": s.get("shift", "day"), "shortage": total,
                        "assigned": s.get("total_assigned", 0), "required": s.get("total_required", 0)})
    return {
        "top_short_details": sorted(by_detail.items(), key=lambda x: -x[1])[:10],
        "top_short_lines": sorted(by_line.items(), key=lambda x: -x[1])[:10],
        "history": by_date,
    }


@api_router.get("/export/{date}")
async def export_schedule(date: str, shift: str = "day"):
    sched_doc = await db.schedules.find_one({"date": date, "shift": shift}, {"_id": 0})
    if not sched_doc:
        raise HTTPException(404, "No schedule")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Schedule {date}"

    assignments = sched_doc["assignments"]
    # Column keys in configured order
    line_configs = sorted(sched_doc["line_configs"], key=lambda c: (c["priority"], c["line"]))
    col_keys: List[str] = []
    for cfg in line_configs:
        for r in range(1, max(1, cfg["run_count"]) + 1):
            col_keys.append(cfg["line"] if r == 1 else f"{cfg['line']} #{r}")

    row_order: List[str] = []
    seen = set()
    for a in assignments:
        rn = a["row_name"]
        if rn not in seen:
            row_order.append(rn)
            seen.add(rn)

    matrix: Dict[Tuple[str, str], dict] = {(a["row_name"], a["line_key"]): a for a in assignments}

    ws.cell(row=1, column=1, value=f"Resource Schedule – {date} ({shift})").font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_keys) + 1)
    ws.cell(row=3, column=1, value="Row Name").font = Font(bold=True)
    for ci, ck in enumerate(col_keys, start=2):
        ws.cell(row=3, column=ci, value=ck).font = Font(bold=True)

    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    red_fill = PatternFill(start_color="FFE3E3", end_color="FFE3E3", fill_type="solid")

    for ri, rn in enumerate(row_order, start=4):
        ws.cell(row=ri, column=1, value=rn).font = Font(bold=True)
        ws.cell(row=ri, column=1).border = border
        for ci, ck in enumerate(col_keys, start=2):
            cell = ws.cell(row=ri, column=ci)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            a = matrix.get((rn, ck))
            if a:
                text = "\n".join(a["assigned_person_names"])
                if a["shortage"] > 0:
                    text += f"\n⚠ SHORT BY {a['shortage']}"
                    cell.fill = red_fill
                cell.value = text
            else:
                cell.value = ""

    # Absent row
    absent_row = len(row_order) + 5
    ws.cell(row=absent_row, column=1, value="ABSENT").font = Font(bold=True, color="CC0000")
    absent_ids = set(sched_doc.get("absent_person_ids", []))
    if absent_ids:
        person_docs = await db.persons.find({"id": {"$in": list(absent_ids)}}, {"_id": 0}).to_list(1000)
        names = ", ".join(f"{p['name']} {p.get('surname','')}".strip() for p in person_docs)
    else:
        names = "None"
    ws.cell(row=absent_row, column=2, value=names)
    ws.merge_cells(start_row=absent_row, start_column=2, end_row=absent_row, end_column=len(col_keys) + 1)

    ws.column_dimensions["A"].width = 26
    for ci in range(2, len(col_keys) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="schedule-{date}-{shift}.xlsx"'},
    )


@api_router.get("/stats")
async def stats():
    persons = await db.persons.count_documents({})
    details = await db.line_details.count_documents({})
    lines = await db.line_details.distinct("line")
    schedules = await db.schedules.count_documents({})
    return {"persons": persons, "details": details, "lines": len(lines), "schedules": schedules}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


@app.on_event("startup")
async def on_startup():
    # Force re-seed if row_name column not yet populated
    try:
        any_detail = await db.line_details.find_one({}, {"_id": 0})
        if any_detail and not any_detail.get("row_name"):
            await db.persons.delete_many({})
            await db.line_details.delete_many({})
            await db.schedules.delete_many({})
            logger.info("Old data lacked row_name; wiped for re-seed")
        await seed_from_file_if_empty()
    except Exception as e:
        logger.exception(f"Seeding failed: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
